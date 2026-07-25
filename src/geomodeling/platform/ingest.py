"""Bounded inspection and immutable standardization of uploaded datasets.

Inspection is read-only and bounded (first 20 rows, full row count, sheet
names, inferred primitive types, candidate mapping). Standardization
converts the mapped columns with ``pandas.to_numeric(errors="coerce")``,
keeps every source row (invalid rows are flagged, never silently dropped),
and writes the Parquet artifact through a temporary file that is read back
and verified before an atomic replace.
"""

from __future__ import annotations

import hashlib
import os
import tempfile
from pathlib import Path
from typing import Any

import pandas as pd

from geomodeling.platform.errors import PlatformError
from geomodeling.platform.schemas import Dimension, FieldMapping
from geomodeling.platform.settings import PlatformSettings

DATASET_TOO_MANY_ROWS = "DATASET_TOO_MANY_ROWS"
SHEET_SELECTION_REQUIRED = "SHEET_SELECTION_REQUIRED"
SHEET_NOT_FOUND = "SHEET_NOT_FOUND"
MAPPING_COLUMN_NOT_FOUND = "MAPPING_COLUMN_NOT_FOUND"
GEOGRAPHIC_NOT_PROJECTED = "GEOGRAPHIC_NOT_PROJECTED"
DATASET_PARSE_FAILED = "DATASET_PARSE_FAILED"

PREVIEW_ROWS = 20

# 常见列名 → 角色候选（仅启发式建议，不自动应用）
_CANDIDATES: dict[str, tuple[str, ...]] = {
    "x": ("x", "easting", "east", "lon", "longitude", "经度"),
    "y": ("y", "northing", "north", "lat", "latitude", "纬度"),
    "z": ("z", "depth", "elevation", "alt", "height", "深度", "高程"),
    "value": ("value", "rho", "val", "v", "grade", "vx", "属性", "值"),
}

STANDARDIZED_SCHEMA = ["source_row", "x", "y", "z", "value", "is_numeric_valid"]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_csv(source_path: Path) -> pd.DataFrame:
    try:
        return pd.read_csv(source_path, encoding="utf-8-sig")
    except (UnicodeDecodeError, pd.errors.ParserError) as exc:
        raise PlatformError(
            DATASET_PARSE_FAILED,
            f"CSV 解析失败：{exc}",
            http_status=400,
        ) from exc


def _xlsx_sheet_names(source_path: Path) -> list[str]:
    from openpyxl import load_workbook

    book = load_workbook(source_path, read_only=True, data_only=True)
    try:
        return list(book.sheetnames)
    finally:
        book.close()


def _read_xlsx(source_path: Path, sheet: str | None) -> tuple[pd.DataFrame, str]:
    names = _xlsx_sheet_names(source_path)
    if sheet is None:
        if len(names) > 1:
            raise PlatformError(
                SHEET_SELECTION_REQUIRED,
                "工作簿包含多个工作表，必须先选择一个",
                {"sheets": names},
            )
        sheet = names[0]
    if sheet not in names:
        raise PlatformError(
            SHEET_NOT_FOUND,
            f"工作表不存在：{sheet}",
            {"sheets": names},
        )
    try:
        frame = pd.read_excel(source_path, sheet_name=sheet, engine="openpyxl")
    except Exception as exc:
        raise PlatformError(DATASET_PARSE_FAILED, f"XLSX 解析失败：{exc}") from exc
    return frame, sheet


def read_source(source_path: Path, suffix: str, sheet: str | None = None) -> tuple[pd.DataFrame, str | None]:
    """Read the uploaded table; returns (frame, resolved_sheet_or_None)."""

    if suffix == "csv":
        return _read_csv(source_path), None
    if suffix == "xlsx":
        return _read_xlsx(source_path, sheet)
    raise PlatformError(
        DATASET_PARSE_FAILED, f"不支持解析的格式：.{suffix}", {"suffix": suffix}
    )


def _check_row_limit(frame: pd.DataFrame, settings: PlatformSettings) -> None:
    if len(frame) > settings.max_upload_rows:
        raise PlatformError(
            DATASET_TOO_MANY_ROWS,
            f"数据行数 {len(frame)} 超过上限 {settings.max_upload_rows}",
            {"row_count": len(frame), "max_upload_rows": settings.max_upload_rows},
        )


def _infer_type(series: pd.Series) -> str:
    if pd.api.types.is_numeric_dtype(series):
        return "numeric"
    converted = pd.to_numeric(series, errors="coerce")
    if converted.notna().all() and series.astype(str).str.strip().ne("").all():
        return "numeric"
    return "text"


def _candidate_mapping(columns: list[str]) -> dict[str, str | None]:
    lowered = {column.lower(): column for column in columns}
    result: dict[str, str | None] = {}
    for role, names in _CANDIDATES.items():
        result[role] = next((lowered[name] for name in names if name in lowered), None)
    return result


def inspect_source(
    settings: PlatformSettings,
    source_path: Path,
    suffix: str,
    sheet: str | None = None,
) -> dict[str, Any]:
    """Bounded inspection of an uploaded source (read-only)."""

    frame, resolved_sheet = read_source(source_path, suffix, sheet)
    _check_row_limit(frame, settings)
    columns = [
        {"name": str(name), "inferred_type": _infer_type(frame[name])}
        for name in frame.columns
    ]
    preview = frame.head(PREVIEW_ROWS).where(pd.notna(frame), None)
    result: dict[str, Any] = {
        "suffix": suffix,
        "sheet": resolved_sheet,
        "columns": columns,
        "preview_rows": preview.to_dict(orient="records"),
        "row_count": len(frame),
        "candidate_mapping": _candidate_mapping([c["name"] for c in columns]),
        "limits": {
            "max_upload_bytes": settings.max_upload_bytes,
            "max_upload_rows": settings.max_upload_rows,
        },
    }
    if suffix == "xlsx":
        result["sheets"] = _xlsx_sheet_names(source_path)
    return result


def standardize(
    settings: PlatformSettings,
    case_id: str,
    dataset_id: str,
    source_path: Path,
    suffix: str,
    mapping: FieldMapping,
    sheet: str | None = None,
) -> dict[str, Any]:
    """Convert the mapped source into the immutable standardized Parquet."""

    import numpy as np

    if mapping.coordinate_kind == "geographic":
        raise PlatformError(
            GEOGRAPHIC_NOT_PROJECTED,
            "经纬度坐标必须先完成投影转换才能建模；v0.4 暂不支持在上游直接以角度插值",
            {"coordinate_kind": mapping.coordinate_kind},
        )

    frame, resolved_sheet = read_source(source_path, suffix, sheet)
    _check_row_limit(frame, settings)

    required = [mapping.x, mapping.y, mapping.value] + ([mapping.z] if mapping.z else [])
    missing = [name for name in required if name not in frame.columns]
    if missing:
        raise PlatformError(
            MAPPING_COLUMN_NOT_FOUND,
            f"映射列在源表中不存在：{missing}",
            {"missing": missing, "columns": [str(c) for c in frame.columns]},
        )

    standardized = pd.DataFrame(
        {
            "source_row": pd.RangeIndex(start=1, stop=len(frame) + 1, step=1),
            "x": pd.to_numeric(frame[mapping.x], errors="coerce"),
            "y": pd.to_numeric(frame[mapping.y], errors="coerce"),
            "z": (
                pd.to_numeric(frame[mapping.z], errors="coerce")
                if mapping.z
                else pd.Series([float("nan")] * len(frame), dtype="float64")
            ),
            "value": pd.to_numeric(frame[mapping.value], errors="coerce"),
        }
    )
    required_numeric = ["x", "y", "value"] + (["z"] if mapping.z else [])
    numeric_block = standardized[required_numeric].to_numpy(dtype="float64", na_value=np.nan)
    standardized["is_numeric_valid"] = np.isfinite(numeric_block).all(axis=1)
    standardized = standardized[STANDARDIZED_SCHEMA]

    target = settings.standardized_dataset(case_id, dataset_id)
    target.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix="standardized-", suffix=".parquet", dir=target.parent)
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        standardized.to_parquet(tmp_path, index=False)
        reread = pd.read_parquet(tmp_path)
        if list(reread.columns) != STANDARDIZED_SCHEMA:
            raise PlatformError(DATASET_PARSE_FAILED, "标准化产物 schema 校验失败")
        if len(reread) != len(standardized):
            raise PlatformError(DATASET_PARSE_FAILED, "标准化产物行数校验失败")
        os.replace(tmp_path, target)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise

    valid_count = int(standardized["is_numeric_valid"].sum())
    return {
        "row_count": len(standardized),
        "valid_row_count": valid_count,
        "invalid_row_count": len(standardized) - valid_count,
        "standardized_path": str(target),
        "standardized_sha256": _sha256(target),
        "sheet": resolved_sheet,
    }
